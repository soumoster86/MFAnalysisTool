"""Polished Excel workbook via openpyxl.

Numbers are written as numbers with cell formats applied, never as pre-rendered
strings — otherwise the recipient cannot sort, total or chart them, which is
the only reason to want a spreadsheet rather than a PDF.

Charts are native Excel charts for the same reason: they stay live and editable.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from services.reports import branding as B
from services.reports.report_data import HOLDING_COLUMNS, ReportData

MONEY = '₹#,##0;[Red]-₹#,##0'
MONEY_SIGNED = '₹#,##0;[Red]-₹#,##0'
PERCENT = '0.0%'
PERCENT_POINTS = '0.0"%"'
NUMBER = '#,##0.000'
RATIO = '0.00'

_THIN = Side(style="thin", color=B.RULE)
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=B.INK)
SUB_FONT = Font(name="Calibri", size=9, color=B.MUTED)
HEAD_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor=B.INK)
BAND_FILL = PatternFill("solid", fgColor=B.BAND)
WARN_FILL = PatternFill("solid", fgColor=B.WARNING_FILL)
WARN_FONT = Font(name="Calibri", size=10, bold=True, color=B.WARNING_INK)
LABEL_FONT = Font(name="Calibri", size=9, color=B.MUTED)
VALUE_FONT = Font(name="Calibri", size=12, bold=True, color=B.INK)
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color=B.INK)

# Cell format per column kind from the shared report model.
KIND_FORMAT = {
    "money": MONEY,
    "percent": PERCENT,
    "percent_points": PERCENT_POINTS,
    "number": NUMBER,
    "ratio": RATIO,
    "text": None,
}


def _title_block(ws: Worksheet, data: ReportData, width: int) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    cell = ws.cell(row=1, column=1, value=data.title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    sub = ws.cell(row=2, column=1, value=data.subtitle)
    sub.font = SUB_FONT

    row = 3
    if data.has_warning:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
        warn = ws.cell(row=row, column=1, value=f"Data quality notice — {data.data_warning}")
        warn.fill = WARN_FILL
        warn.font = WARN_FONT
        warn.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row].height = 42
        row += 1
    return row + 1


def _section(ws: Worksheet, row: int, text: str, width: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = BAND_FILL
    return row + 1


def _autosize(ws: Worksheet, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _summary_sheet(wb: Workbook, data: ReportData) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    row = _title_block(ws, data, width=6)

    # KPI cards, three across.
    row = _section(ws, row, "Portfolio at a glance", 6)
    start = row + 1
    for i, kpi in enumerate(data.kpis):
        col = 1 + (i % 3) * 2
        line = start + (i // 3) * 3
        label = ws.cell(row=line, column=col, value=kpi.label)
        label.font = LABEL_FONT
        value = ws.cell(row=line + 1, column=col, value=kpi.value)
        value.font = Font(
            name="Calibri", size=12, bold=True, color=B.tone_colour(kpi.tone)
        )
        if kpi.caption:
            cap = ws.cell(row=line + 2, column=col, value=kpi.caption)
            cap.font = LABEL_FONT
    row = start + ((len(data.kpis) + 2) // 3) * 3 + 1

    if data.risk.rows:
        row = _section(ws, row, data.risk.title, 6)
        for label, value in data.risk.rows:
            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
            ws.cell(row=row, column=2, value=value).font = Font(bold=True)
            row += 1
        note = ws.cell(row=row, column=1, value=data.risk.note)
        note.font = SUB_FONT
        row += 2

    if data.overlap_rows:
        row = _section(ws, row, "Fund overlap", 6)
        for label, value in data.overlap_rows:
            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
            ws.cell(row=row, column=2, value=value).font = Font(bold=True)
            row += 1
        row += 1

    if data.ai_summary:
        row = _section(ws, row, "AI review", 6)
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 6, end_column=6)
        cell = ws.cell(row=row, column=1, value=data.ai_summary)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    _autosize(ws, {1: 34, 2: 20, 3: 34, 4: 20, 5: 34, 6: 20})


def _holdings_sheet(wb: Workbook, data: ReportData) -> None:
    if not data.holdings:
        return
    ws = wb.create_sheet("Holdings")
    ws.sheet_view.showGridLines = False
    row = _title_block(ws, data, width=len(HOLDING_COLUMNS))

    header_row = row
    for col, (_, label, _) in enumerate(HOLDING_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[header_row].height = 22

    for i, record in enumerate(data.holdings):
        line = header_row + 1 + i
        for col, (key, _, kind) in enumerate(HOLDING_COLUMNS, start=1):
            value = record.get(key)
            # Weight is stored in percentage points but Excel's % format
            # multiplies by 100, so scale it to a fraction for the cell.
            if kind == "percent_points" and value is not None:
                cell = ws.cell(row=line, column=col, value=float(value) / 100)
                cell.number_format = PERCENT
            else:
                cell = ws.cell(row=line, column=col, value=value)
                fmt = KIND_FORMAT.get(kind)
                if fmt:
                    cell.number_format = fmt
            cell.border = BORDER
            if i % 2:
                cell.fill = BAND_FILL
            if kind != "text":
                cell.alignment = Alignment(horizontal="right")

    last = header_row + len(data.holdings)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(HOLDING_COLUMNS))}{last}"
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Totals, so the sheet reconciles without the reader adding it up.
    total_row = last + 1
    ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
    for col, (key, _, kind) in enumerate(HOLDING_COLUMNS, start=1):
        if key in {"invested_amount", "current_value", "gain"}:
            letter = get_column_letter(col)
            cell = ws.cell(
                row=total_row, column=col,
                value=f"=SUM({letter}{header_row + 1}:{letter}{last})",
            )
            cell.number_format = MONEY
            cell.font = Font(bold=True)
            cell.border = BORDER

    _autosize(ws, {1: 46, 2: 16, 3: 13, 4: 15, 5: 15, 6: 15, 7: 11, 8: 10})


def _allocation_sheet(wb: Workbook, data: ReportData) -> None:
    if not data.allocations:
        return
    ws = wb.create_sheet("Allocation")
    ws.sheet_view.showGridLines = False
    row = _title_block(ws, data, width=4)

    for name, mapping in data.allocations.items():
        row = _section(ws, row, name, 4)
        head = row
        ws.cell(row=head, column=1, value="Bucket").font = HEAD_FONT
        ws.cell(row=head, column=1).fill = HEAD_FILL
        ws.cell(row=head, column=2, value="Weight").font = HEAD_FONT
        ws.cell(row=head, column=2).fill = HEAD_FILL
        for i, (bucket, weight) in enumerate(mapping.items(), start=1):
            ws.cell(row=head + i, column=1, value=bucket).border = BORDER
            cell = ws.cell(row=head + i, column=2, value=float(weight) / 100)
            cell.number_format = PERCENT
            cell.border = BORDER

        end = head + len(mapping)
        chart = PieChart()
        chart.title = name
        chart.height, chart.width = 7.5, 12
        chart.add_data(Reference(ws, min_col=2, min_row=head, max_row=end), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=head + 1, max_row=end))
        ws.add_chart(chart, f"E{head}")
        row = end + 2

    if data.top_holdings:
        row = _section(ws, row, "Largest look-through holdings", 4)
        head = row
        ws.cell(row=head, column=1, value="Security").font = HEAD_FONT
        ws.cell(row=head, column=1).fill = HEAD_FILL
        ws.cell(row=head, column=2, value="Weight").font = HEAD_FONT
        ws.cell(row=head, column=2).fill = HEAD_FILL
        for i, (security, weight) in enumerate(data.top_holdings, start=1):
            ws.cell(row=head + i, column=1, value=security).border = BORDER
            cell = ws.cell(row=head + i, column=2, value=float(weight) / 100)
            cell.number_format = PERCENT
            cell.border = BORDER
        end = head + len(data.top_holdings)

        bar = BarChart()
        bar.type = "bar"
        bar.title = "Look-through stock weights"
        bar.height, bar.width = 9, 12
        bar.add_data(Reference(ws, min_col=2, min_row=head, max_row=end), titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1, min_row=head + 1, max_row=end))
        bar.legend = None
        ws.add_chart(bar, f"E{head}")

    _autosize(ws, {1: 40, 2: 14, 3: 4, 4: 4})


def _sources_sheet(wb: Workbook, data: ReportData) -> None:
    if not data.source_rows and not data.notes:
        return
    ws = wb.create_sheet("Data sources")
    ws.sheet_view.showGridLines = False
    row = _title_block(ws, data, width=2)

    if data.source_rows:
        row = _section(ws, row, "Where each figure came from", 2)
        ws.cell(row=row, column=1, value="Fund").font = HEAD_FONT
        ws.cell(row=row, column=1).fill = HEAD_FILL
        ws.cell(row=row, column=2, value="Source").font = HEAD_FONT
        ws.cell(row=row, column=2).fill = HEAD_FILL
        row += 1
        for label, source in data.source_rows:
            ws.cell(row=row, column=1, value=label).border = BORDER
            cell = ws.cell(row=row, column=2, value=source)
            cell.border = BORDER
            # Fabricated inputs must stand out in the audit trail too.
            if "synthetic" in source.lower() or "sample" in source.lower():
                cell.font = Font(bold=True, color=B.NEGATIVE)
            row += 1
        row += 1

    if data.notes:
        row = _section(ws, row, "Analysis notes", 2)
        for note in data.notes:
            cell = ws.cell(row=row, column=1, value=note)
            cell.alignment = Alignment(wrap_text=True)
            row += 1

    _autosize(ws, {1: 62, 2: 28})


def build_workbook(data: ReportData) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    _summary_sheet(wb, data)
    _holdings_sheet(wb, data)
    _allocation_sheet(wb, data)
    _sources_sheet(wb, data)

    wb.properties.title = data.title
    wb.properties.creator = "MF Analysis Tool"
    wb.properties.description = B.FOOTER_NOTE
    return wb


def excel_bytes(data: ReportData) -> bytes:
    buffer = BytesIO()
    build_workbook(data).save(buffer)
    return buffer.getvalue()
